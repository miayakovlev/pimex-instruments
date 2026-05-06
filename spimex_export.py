#!/usr/bin/env python3
"""
Собирает с карточек инструментов SPIMEX (нефтепродукты):
- наименование из поля «Биржевой товар»;
- текст из поля «Базис поставки»;
- значение «Последняя» для самой поздней даты в таблице
  «Результаты последних 10 торговых сессий».

Выход: CSV (UTF-8 с BOM для Excel).

Использование данных сайта биржи — на условиях пользовательского соглашения SPIMEX;
при массовой автоматизации и коммерческом распространении сверьтесь с правилами и,
при необходимости, с официальными каналами предоставления информации.
"""

from __future__ import annotations

import argparse
import csv
import os
import re
import smtplib
import ssl
import sys
import time
from dataclasses import dataclass
from datetime import datetime
from email.message import EmailMessage
from pathlib import Path
from urllib.parse import parse_qs, urlparse
from zoneinfo import ZoneInfo

try:
    import matplotlib.pyplot as plt
except ImportError:
    plt = None
try:
    from openpyxl import Workbook
    from openpyxl.chart import LineChart, Reference
    from openpyxl.utils import get_column_letter
except ImportError:
    Workbook = None
import requests
from bs4 import BeautifulSoup
from requests.exceptions import ProxyError

SESSION = requests.Session()
SESSION.headers.update(
    {
        "User-Agent": (
            "Mozilla/5.0 (Macintosh; Intel Mac OS X 10_15_7) "
            "AppleWebKit/537.36 (KHTML, like Gecko) Chrome/120.0.0.0 Safari/537.36"
        ),
        "Accept-Language": "ru-RU,ru;q=0.9,en;q=0.8",
    }
)

DATE_RE = re.compile(r"^\d{1,2}\.\d{1,2}\.\d{4}$")
BIRZHEVOI_Tovar_RE = re.compile(r"биржев(ой|ый)\s+товар", re.IGNORECASE)
BAZIS_POSTAVKI_RE = re.compile(r"базис\s+поставки", re.IGNORECASE)
SESSIONS_SECTION_RE = re.compile(
    r"Результаты последних\s+10\s+торговых\s+сессий", re.IGNORECASE
)
DETAIL_URL = (
    "https://spimex.com/markets/oil_products/instruments/list/detail.php?code="
)


def normalize_ws(s: str) -> str:
    return re.sub(r"\s+", " ", s.replace("\xa0", " ")).strip()


def instrument_code(url_or_code: str) -> str:
    s = url_or_code.strip()
    if "://" in s or s.startswith("//"):
        q = urlparse(s).query
        parts = parse_qs(q).get("code")
        if parts and parts[0]:
            return parts[0].strip()
        raise ValueError(f"В URL не найден параметр code: {s}")
    return s.strip()


def fetch_html(code: str) -> str:
    url = DETAIL_URL + requests.utils.quote(code, safe="")
    try:
        r = SESSION.get(url, timeout=45)
        r.raise_for_status()
    except ProxyError:
        # Часто в cron/systemd подхватывается некорректный env-proxy:
        # повторяем запрос напрямую без переменных окружения прокси.
        direct_session = requests.Session()
        direct_session.trust_env = False
        direct_session.headers.update(SESSION.headers)
        r = direct_session.get(url, timeout=45)
        r.raise_for_status()
    # Частые кракозябры без указания кодировки
    if r.encoding is None or r.encoding.lower() == "iso-8859-1":
        r.encoding = r.apparent_encoding or "utf-8"
    return r.text


def fetch_html_playwright(code: str) -> str:
    """Страница может подставлять таблицы через JS — тогда нужен браузер."""
    try:
        from playwright.sync_api import sync_playwright
    except ImportError as e:
        raise RuntimeError(
            "Нужен Playwright: pip install playwright && playwright install chromium"
        ) from e

    url = DETAIL_URL + requests.utils.quote(code, safe="")
    with sync_playwright() as p:
        browser = p.chromium.launch(headless=True)
        try:
            page = browser.new_page()
            page.goto(url, wait_until="domcontentloaded", timeout=90_000)
            page.wait_for_timeout(4000)
            return page.content()
        finally:
            browser.close()


def cell_text(cell) -> str:
    return cell.get_text(" ", strip=True)


def rows_cells(table):
    """Списки текста ячеек по каждой строке (без вложенных таблиц в ячейке — редкость)."""
    out = []
    for tr in table.find_all("tr"):
        cells = tr.find_all(["td", "th"], recursive=False)
        if not cells:
            continue
        out.append([cell_text(c) for c in cells])
    return out


def extract_product_desc_cell(soup: BeautifulSoup, label_re: re.Pattern[str]) -> str | None:
    """Строка таблицы описания инструмента: первая ячейка — подпись, дальше код/ссылка и полное имя."""
    for tr in soup.find_all("tr"):
        cells = tr.find_all(["td", "th"], recursive=False)
        if not cells:
            continue
        first = normalize_ws(cell_text(cells[0]))
        if not label_re.search(first):
            continue
        rest = [normalize_ws(cell_text(c)) for c in cells[1:] if normalize_ws(cell_text(c))]
        if not rest:
            return None
        return rest[-1]
    return None


def extract_birzhevoi_tovar(soup: BeautifulSoup) -> str | None:
    return extract_product_desc_cell(soup, BIRZHEVOI_Tovar_RE)


def extract_bazis_postavki(soup: BeautifulSoup) -> str | None:
    return extract_product_desc_cell(soup, BAZIS_POSTAVKI_RE)


def find_sessions_table(soup: BeautifulSoup):
    """
    Таблица последних торгов имеет thead с «Дата торгов», «Последняя», а не текст заголовка
    раздела (он обычно снаружи). Нельзя брать просто «следующую table» после заголовка —
    часто между ними идёт product_desc или другая разметка.
    """
    for table in soup.find_all("table"):
        tn = normalize_ws(table.get_text(" ", strip=True)).lower()
        if "последняя" not in tn or "дата торгов" not in tn:
            continue
        rows = rows_cells(table)
        if any(r for r in rows if r and DATE_RE.match(r[0].strip())):
            return table

    # Дополнительно: класс с вёрстки SPIMEX
    for table in soup.select("table.results_table"):
        tn = normalize_ws(table.get_text(" ", strip=True)).lower()
        if "последняя" in tn and rows_cells(table):
            rows = rows_cells(table)
            if any(r for r in rows if r and DATE_RE.match(r[0].strip())):
                return table

    marker_sub = "последних 10 торговых сессий"
    marker_full = "Результаты последних 10 торговых сессий"
    for tag in soup.find_all(["h2", "h3", "h4", "div", "p", "strong", "td", "th"]):
        head = normalize_ws(tag.get_text(" ", strip=True))
        if marker_full.lower() not in head.lower() and marker_sub.lower() not in head.lower():
            continue
        for table in tag.find_all_next("table", limit=8):
            tn = normalize_ws(table.get_text(" ", strip=True)).lower()
            if "последняя" in tn and "дата торгов" in tn:
                rows = rows_cells(table)
                if any(r for r in rows if r and DATE_RE.match(r[0].strip())):
                    return table
    return None


def column_index_poslednyaya(header_rows: list[list[str]]) -> int | None:
    """
    Индекс колонки «Последняя» в строках данных (td).

    У SPIMEX первая строка thead: «Дата торгов» (rowspan=2) + «Цена, руб.» (colspan=3) + объёмы;
    вторая — только «Рыночная | Первая | Последняя» без ячейки даты. Тогда индекс в подстроке цен
    нужно сдвинуть на +1 относительно второй строки заголовка.
    """
    price_head = {"рыночная", "первая", "последняя"}
    for row in header_rows:
        if not row:
            continue
        lowered = [normalize_ws(h).lower() for h in row]
        if "последняя" not in lowered:
            continue
        idx = lowered.index("последняя")
        if len(row) == 3 and all(h in price_head for h in lowered):
            return idx + 1
        return idx
    return None


def parse_sessions_last_price(table) -> tuple[str | None, str | None]:
    """
    Возвращает (дата DD.MM.YYYY, значение «Последняя») для самой поздней даты в таблице.
    """
    prices = parse_sessions_prices(table)
    if not prices:
        return None, None
    latest_date = max(prices.keys(), key=lambda x: datetime.strptime(x, "%d.%m.%Y"))
    return latest_date, prices[latest_date]


def parse_sessions_prices(table) -> dict[str, str]:
    """
    Возвращает словарь {дата DD.MM.YYYY: значение «Последняя»} по всем строкам таблицы.
    """
    rows = rows_cells(table)
    if not rows:
        return {}

    # Заголовок: 1–2 верхние строки, пока не встретим строку с датой в первом столбце.
    header_rows: list[list[str]] = []
    data_start = 0
    for i, row in enumerate(rows):
        if row and DATE_RE.match(row[0].strip()):
            data_start = i
            break
        header_rows.append(row)
        data_start = i + 1

    last_idx = column_index_poslednyaya(header_rows)
    if last_idx is None:
        # Типичная вёрстка SPIMEX: дата + три цены + объёмы (см. пример на сайте).
        last_idx = 3

    out: dict[str, str] = {}

    for row in rows[data_start:]:
        if not row:
            continue
        d0 = row[0].strip()
        if not DATE_RE.match(d0):
            continue
        if last_idx >= len(row):
            continue
        try:
            d = datetime.strptime(d0, "%d.%m.%Y")
        except ValueError:
            continue
        val = row[last_idx].strip()
        if val:
            out[d0] = normalize_ws(val)

    return out


@dataclass
class Row:
    code: str
    url: str
    birzhevoi_tovar: str | None
    bazis_postavki: str | None
    data_posledney_sessii: str | None
    tsena_poslednyaya: str | None
    sessions_last_by_date: dict[str, str]


BASE_FIELDNAMES = [
    "code",
    "url",
    "birzhevoi_tovar",
    "bazis_postavki",
]


def extract_one(code: str, *, use_browser: bool) -> Row:
    url = DETAIL_URL + requests.utils.quote(code, safe="")
    if use_browser:
        html = fetch_html_playwright(code)
    else:
        html = fetch_html(code)
    soup = BeautifulSoup(html, "html.parser")
    commodity = extract_birzhevoi_tovar(soup)
    bazis = extract_bazis_postavki(soup)
    sess_table = find_sessions_table(soup)
    sess_date = sess_last = None
    sessions_prices: dict[str, str] = {}
    if sess_table:
        sessions_prices = parse_sessions_prices(sess_table)
        if sessions_prices:
            sess_date = max(sessions_prices.keys(), key=lambda x: datetime.strptime(x, "%d.%m.%Y"))
            sess_last = sessions_prices[sess_date]

    # Пустая разметка в requests-трафике: один раз пробуем через браузер, если включено авто.
    return Row(
        code=code,
        url=url,
        birzhevoi_tovar=commodity,
        bazis_postavki=bazis,
        data_posledney_sessii=sess_date,
        tsena_poslednyaya=sess_last,
        sessions_last_by_date=sessions_prices,
    )


def _today_msk_date_str() -> str:
    return datetime.now(ZoneInfo("Europe/Moscow")).strftime("%d.%m.%Y")


def _is_date_column(name: str) -> bool:
    return bool(DATE_RE.match(name.strip()))


def _parse_price_value(raw: str) -> float | None:
    s = normalize_ws(raw).replace(" ", "").replace(",", ".")
    if not s:
        return None
    try:
        return float(s)
    except ValueError:
        return None


def _format_price_value(value: str | None) -> str:
    if not value:
        return ""
    return normalize_ws(value)


def _read_existing_wide_csv(path: Path) -> tuple[list[str], dict[str, dict[str, str]]]:
    if not path.exists():
        return [], {}
    with open(path, encoding="utf-8-sig", newline="") as fp:
        reader = csv.DictReader(fp)
        if not reader.fieldnames:
            return [], {}
        date_columns = [c for c in reader.fieldnames if _is_date_column(c)]
        data: dict[str, dict[str, str]] = {}
        for row in reader:
            code = (row.get("code") or "").strip()
            if not code:
                continue
            normalized: dict[str, str] = {}
            for key, value in row.items():
                if key is None:
                    continue
                normalized[key] = (value or "").strip()
            data[code] = normalized
    return date_columns, data


def _sorted_date_columns(cols: list[str]) -> list[str]:
    uniq = sorted(set(cols), key=lambda x: datetime.strptime(x, "%d.%m.%Y"))
    return uniq


def _latest_known_daily_price(prev: dict[str, str], date_cols: list[str]) -> str:
    for d in reversed(date_cols):
        v = (prev.get(d) or "").strip()
        if v:
            return v
    return ""


def _build_wide_rows(
    fresh_rows: list[Row],
    existing_rows: dict[str, dict[str, str]],
    date_columns: list[str],
    run_date: str,
) -> tuple[list[str], list[dict[str, str]]]:
    all_date_columns = _sorted_date_columns(date_columns + [run_date])
    rows_for_csv: list[dict[str, str]] = []

    for r in fresh_rows:
        prev = existing_rows.get(r.code, {})
        merged: dict[str, str] = {}
        merged["code"] = r.code
        merged["url"] = r.url
        merged["birzhevoi_tovar"] = r.birzhevoi_tovar or prev.get("birzhevoi_tovar", "")
        merged["bazis_postavki"] = r.bazis_postavki or prev.get("bazis_postavki", "")

        for d in all_date_columns:
            merged[d] = prev.get(d, "")

        # Колонка даты запуска: актуальная/последняя доступная с сайта; при ошибке — последняя известная локально.
        merged[run_date] = _format_price_value(r.tsena_poslednyaya) or _latest_known_daily_price(
            prev,
            all_date_columns,
        )
        rows_for_csv.append(merged)

    fieldnames = BASE_FIELDNAMES + all_date_columns
    return fieldnames, rows_for_csv


def _write_wide_csv(path: Path, fieldnames: list[str], rows_for_csv: list[dict[str, str]]) -> None:
    with open(path, "w", encoding="utf-8-sig", newline="") as fp:
        writer = csv.DictWriter(fp, fieldnames=fieldnames)
        writer.writeheader()
        for row in rows_for_csv:
            writer.writerow(row)


def _is_error_text(value: str | None) -> bool:
    s = (value or "").strip()
    return s.startswith("REQUEST_ERROR:") or s.startswith("RUNTIME_ERROR:")


def _read_history_csv(path: Path) -> dict[str, dict[str, str | dict[str, str]]]:
    if not path.exists():
        return {}
    out: dict[str, dict[str, str | dict[str, str]]] = {}
    with open(path, encoding="utf-8-sig", newline="") as fp:
        reader = csv.DictReader(fp)
        for row in reader:
            code = (row.get("code") or "").strip()
            trade_date = (row.get("trade_date") or "").strip()
            price_last = normalize_ws((row.get("price_last") or "").strip())
            if not code or not trade_date or not price_last:
                continue
            rec = out.setdefault(
                code,
                {
                    "url": (row.get("url") or "").strip(),
                    "birzhevoi_tovar": (row.get("birzhevoi_tovar") or "").strip(),
                    "bazis_postavki": (row.get("bazis_postavki") or "").strip(),
                    "prices": {},
                },
            )
            prices = rec["prices"]
            if isinstance(prices, dict):
                prices[trade_date] = price_last
    return out


def _merge_history(
    history: dict[str, dict[str, str | dict[str, str]]],
    fresh_rows: list[Row],
) -> dict[str, dict[str, str | dict[str, str]]]:
    for row in fresh_rows:
        rec = history.setdefault(
            row.code,
            {"url": row.url, "birzhevoi_tovar": "", "bazis_postavki": "", "prices": {}},
        )
        rec["url"] = row.url
        if row.birzhevoi_tovar and not _is_error_text(row.birzhevoi_tovar):
            rec["birzhevoi_tovar"] = row.birzhevoi_tovar
        if row.bazis_postavki:
            rec["bazis_postavki"] = row.bazis_postavki
        prices = rec["prices"]
        if isinstance(prices, dict):
            for d, v in row.sessions_last_by_date.items():
                if d and v:
                    prices[d] = normalize_ws(v)
    return history


def _write_history_csv(path: Path, history: dict[str, dict[str, str | dict[str, str]]]) -> int:
    fieldnames = ["code", "url", "birzhevoi_tovar", "bazis_postavki", "trade_date", "price_last"]
    rows: list[dict[str, str]] = []
    for code in sorted(history.keys()):
        rec = history[code]
        prices = rec.get("prices", {})
        if not isinstance(prices, dict):
            continue
        for d in sorted(prices.keys(), key=lambda x: datetime.strptime(x, "%d.%m.%Y")):
            rows.append(
                {
                    "code": code,
                    "url": str(rec.get("url", "")),
                    "birzhevoi_tovar": str(rec.get("birzhevoi_tovar", "")),
                    "bazis_postavki": str(rec.get("bazis_postavki", "")),
                    "trade_date": d,
                    "price_last": normalize_ws(str(prices[d])),
                },
            )
    with open(path, "w", encoding="utf-8-sig", newline="") as fp:
        writer = csv.DictWriter(fp, fieldnames=fieldnames)
        writer.writeheader()
        writer.writerows(rows)
    return len(rows)


def _draw_price_histogram(path: Path, date_columns: list[str], rows_for_csv: list[dict[str, str]]) -> bool:
    if plt is None:
        return False
    if not date_columns or not rows_for_csv:
        return False

    data_rows: list[tuple[str, list[float], list[str]]] = []
    for row in rows_for_csv:
        values: list[float] = []
        dates: list[str] = []
        label = row.get("code", "").strip()
        for d in date_columns:
            val = _parse_price_value(row.get(d, ""))
            if val is None:
                continue
            values.append(val)
            dates.append(d)
        if values:
            data_rows.append((label, values, dates))

    if not data_rows:
        return False

    fig_h = max(3.5, 2.6 * len(data_rows))
    fig, axes = plt.subplots(len(data_rows), 1, figsize=(14, fig_h), squeeze=False)
    axes_flat = [ax for sub in axes for ax in sub]

    for ax, (label, values, dates) in zip(axes_flat, data_rows):
        ax.bar(dates, values, color="#4E79A7")
        ax.set_title(label, loc="left", fontsize=10)
        ax.set_ylabel("Цена")
        ax.grid(axis="y", alpha=0.25)
        ax.tick_params(axis="x", labelrotation=45, labelsize=8)

    fig.suptitle("Динамика цены по датам (Последняя)", fontsize=13)
    fig.tight_layout()
    fig.savefig(path, dpi=150)
    plt.close(fig)
    return True


def _build_history_wide_table(
    history: dict[str, dict[str, str | dict[str, str]]],
) -> tuple[list[str], list[dict[str, str]]]:
    all_trade_dates: list[str] = []
    for rec in history.values():
        prices = rec.get("prices", {})
        if isinstance(prices, dict):
            all_trade_dates.extend(prices.keys())
    unique_dates = _sorted_date_columns(list(set(all_trade_dates))) if all_trade_dates else []

    fieldnames = BASE_FIELDNAMES + unique_dates
    rows: list[dict[str, str]] = []
    for code in sorted(history.keys()):
        rec = history[code]
        prices = rec.get("prices", {})
        prices_map = prices if isinstance(prices, dict) else {}
        row = {
            "code": code,
            "url": str(rec.get("url", "")),
            "birzhevoi_tovar": str(rec.get("birzhevoi_tovar", "")),
            "bazis_postavki": str(rec.get("bazis_postavki", "")),
        }
        last_known = ""
        for d in unique_dates:
            raw = normalize_ws(str(prices_map.get(d, "")))
            if raw:
                last_known = raw
                row[d] = raw
            else:
                # Для непрерывного графика: если за дату нет новой сделки, тянем прошлую известную цену.
                row[d] = last_known
        rows.append(row)
    return fieldnames, rows


def _write_xlsx_report(
    path: Path,
    history: dict[str, dict[str, str | dict[str, str]]],
) -> bool:
    if Workbook is None:
        return False
    fieldnames, rows_for_csv = _build_history_wide_table(history)

    wb = Workbook()
    ws_data = wb.active
    ws_data.title = "Отчёт"
    ws_data.append(fieldnames)

    date_cols = [i for i, name in enumerate(fieldnames, start=1) if _is_date_column(name)]
    for row in rows_for_csv:
        values = []
        for idx, key in enumerate(fieldnames, start=1):
            raw = row.get(key, "")
            if idx in date_cols:
                parsed = _parse_price_value(raw)
                values.append(parsed if parsed is not None else None)
            else:
                values.append(raw)
        ws_data.append(values)

    for col in ws_data.columns:
        letter = col[0].column_letter
        max_len = max(len(str(c.value)) if c.value is not None else 0 for c in col)
        ws_data.column_dimensions[letter].width = min(max(10, max_len + 2), 70)

    if date_cols and ws_data.max_row >= 2:
        line_chart = LineChart()
        line_chart.title = "Динамика цены по всем продуктам"
        line_chart.style = 2
        line_chart.y_axis.title = "Цена"
        line_chart.x_axis.title = "Дата"
        line_chart.x_axis.tickLblPos = "low"
        line_chart.y_axis.tickLblPos = "nextTo"
        line_chart.y_axis.numFmt = "# ##0"
        line_chart.x_axis.numFmt = "@"
        line_chart.x_axis.delete = False
        line_chart.y_axis.delete = False
        line_chart.width = 38
        line_chart.height = 18
        line_chart.legend.position = "r"
        line_chart.legend.overlay = False

        data_ref = Reference(
            ws_data,
            min_col=1,
            max_col=date_cols[-1],
            min_row=2,
            max_row=ws_data.max_row,
        )
        categories = Reference(
            ws_data,
            min_col=date_cols[0],
            max_col=date_cols[-1],
            min_row=1,
            max_row=1,
        )
        line_chart.add_data(data_ref, from_rows=True, titles_from_data=True)
        line_chart.set_categories(categories)

        chart_anchor = f"A{ws_data.max_row + 3}"
        ws_data.add_chart(line_chart, chart_anchor)

        # Отдельный лист с малыми графиками по каждому продукту для удобства чтения.
        ws_small = wb.create_sheet("Графики")
        charts_per_row = 2
        row_step = 18
        col_step = 9
        start_row = 1
        start_col = 1
        for idx, data_row in enumerate(range(2, ws_data.max_row + 1)):
            code = ws_data.cell(row=data_row, column=1).value or f"row_{data_row}"
            c = LineChart()
            c.title = str(code)
            c.style = 4
            c.y_axis.title = "Цена"
            c.x_axis.title = "Дата"
            c.y_axis.numFmt = "# ##0"
            c.width = 14
            c.height = 7
            c.legend = None

            row_ref = Reference(
                ws_data,
                min_col=date_cols[0],
                max_col=date_cols[-1],
                min_row=data_row,
                max_row=data_row,
            )
            c.add_data(row_ref, from_rows=True, titles_from_data=False)
            c.set_categories(categories)

            if c.series:
                c.series[0].marker.symbol = "circle"
                c.series[0].marker.size = 5
                c.series[0].graphicalProperties.line.width = 20000

            grid_row = idx // charts_per_row
            grid_col = idx % charts_per_row
            anchor_col = start_col + (grid_col * col_step)
            anchor_row = start_row + (grid_row * row_step)
            ws_small.add_chart(c, f"{get_column_letter(anchor_col)}{anchor_row}")
    else:
        ws_data.cell(row=2, column=len(fieldnames) + 2, value="Недостаточно данных для построения графиков.")

    wb.save(path)
    return True


def _env_bool(key: str, default: bool = False) -> bool:
    v = os.environ.get(key, "").strip().lower()
    if not v:
        return default
    return v in ("1", "true", "yes", "on")


def _mail_recipients_from_env() -> list[str]:
    raw = os.environ.get("SPIMEX_MAIL_TO", "")
    return [x.strip() for x in raw.split(",") if x.strip()]


def send_report_email(
    report_path: Path,
    recipients: list[str],
    *,
    subject: str | None,
) -> None:
    host = os.environ.get("SPIMEX_SMTP_HOST", "").strip()
    mail_from = os.environ.get("SPIMEX_MAIL_FROM", "").strip()
    port = int(os.environ.get("SPIMEX_SMTP_PORT", "587") or "587")
    user = os.environ.get("SPIMEX_SMTP_USER", "").strip() or None
    password = os.environ.get("SPIMEX_SMTP_PASSWORD", "").strip() or None
    use_ssl = _env_bool("SPIMEX_SMTP_USE_SSL", port == 465)

    if not host:
        raise ValueError("Для отправки почты задайте SPIMEX_SMTP_HOST")
    if not mail_from:
        raise ValueError("Задайте SPIMEX_MAIL_FROM (адрес отправителя)")
    if not recipients:
        raise ValueError("Нет получателей письма")

    msk_now = datetime.now(ZoneInfo("Europe/Moscow"))
    subj = subject or f"SPIMEX: выгрузка инструментов {msk_now:%Y-%m-%d %H:%M} МСК"

    msg = EmailMessage()
    msg["Subject"] = subj
    msg["From"] = mail_from
    msg["To"] = ", ".join(recipients)
    msg.set_content(
        f"Выгрузка SPIMEX: файл {report_path.name} во вложении (перезаписан на сервере).\n"
        f"Отправлено {msk_now:%Y-%m-%d %H:%M:%S} МСК.\n"
    )
    data = report_path.read_bytes()
    msg.add_attachment(
        data,
        maintype="application",
        subtype="vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        filename=report_path.name,
    )

    ctx = ssl.create_default_context()
    if use_ssl:
        with smtplib.SMTP_SSL(host, port, timeout=90, context=ctx) as smtp:
            if user and password:
                smtp.login(user, password)
            smtp.send_message(msg)
    else:
        with smtplib.SMTP(host, port, timeout=90) as smtp:
            smtp.starttls(context=ctx)
            if user and password:
                smtp.login(user, password)
            smtp.send_message(msg)


def extract_one_smart(code: str, *, browser: bool, browser_fallback: bool) -> Row:
    if browser:
        return extract_one(code, use_browser=True)
    row = extract_one(code, use_browser=False)
    if browser_fallback and (row.birzhevoi_tovar is None or row.tsena_poslednyaya is None):
        return extract_one(code, use_browser=True)
    return row


def main(argv: list[str]) -> int:
    p = argparse.ArgumentParser(description="Выгрузка полей карточки инструмента SPIMEX в CSV.")
    p.add_argument(
        "items",
        nargs="*",
        help="Код инструмента или полный URL detail.php?code=...",
    )
    p.add_argument(
        "-f",
        "--file",
        metavar="PATH",
        help="Файл: по одному коду или URL в строке; пустые и # игнорируются.",
    )
    p.add_argument(
        "-o",
        "--output",
        default="spimex_instruments.csv",
        help="Путь итогового CSV (wide-формат, UTF-8 с BOM). По умолчанию spimex_instruments.csv",
    )
    p.add_argument(
        "--delay",
        type=float,
        default=0.75,
        help="Пауза между запросами, сек. (умеренная нагрузка на сайт).",
    )
    p.add_argument(
        "--browser",
        action="store_true",
        help="Загружать страницу через Chromium (Playwright); нужен pip install playwright && playwright install chromium.",
    )
    p.add_argument(
        "--browser-fallback",
        action="store_true",
        help="Если после обычного запроса не найдены товар и/или «Последняя», повторить через браузер.",
    )
    p.add_argument(
        "--dump-html",
        metavar="PATH",
        help="Сохранить HTML первого инструмента для отладки и выйти (коды из аргументов/-f).",
    )
    p.add_argument(
        "--mail-to",
        action="append",
        default=[],
        metavar="EMAIL",
        help="Получатель копии CSV (можно несколько раз). Если не задано — SPIMEX_MAIL_TO в окружении.",
    )
    p.add_argument(
        "--mail-subject",
        default=None,
        help="Тема письма (иначе автоматически с датой по Москве).",
    )
    p.add_argument(
        "--histogram-output",
        default="spimex_prices_hist.png",
        metavar="PATH",
        help="Путь PNG-гистограммы динамики цен по продуктам. По умолчанию spimex_prices_hist.png",
    )
    p.add_argument(
        "--report-output",
        default="spimex_report.xlsx",
        metavar="PATH",
        help="Путь Excel-отчёта (таблица + гистограмма в одном файле). По умолчанию spimex_report.xlsx",
    )
    p.add_argument(
        "--history-output",
        default="spimex_price_history.csv",
        metavar="PATH",
        help="Путь CSV с накопленной историей торговых дат и цен. По умолчанию spimex_price_history.csv",
    )
    args = p.parse_args(argv)

    codes: list[str] = []
    if args.file:
        with open(args.file, encoding="utf-8") as fp:
            for line in fp:
                line = line.strip()
                if not line or line.startswith("#"):
                    continue
                codes.append(instrument_code(line))
    for x in args.items:
        codes.append(instrument_code(x))

    if not codes:
        p.error("Укажите коды или URL в аргументах или через --file.")

    if args.dump_html:
        c0 = codes[0]
        try:
            html = fetch_html_playwright(c0) if args.browser else fetch_html(c0)
        except requests.RequestException as e:
            print(f"Ошибка загрузки: {e}", file=sys.stderr)
            return 1
        with open(args.dump_html, "w", encoding="utf-8") as fp:
            fp.write(html)
        print(f"HTML сохранён: {args.dump_html} (код {c0})")
        return 0

    rows_out: list[Row] = []
    request_errors: list[tuple[str, str]] = []
    runtime_errors: list[tuple[str, str]] = []
    for i, c in enumerate(codes):
        if i and args.delay:
            time.sleep(args.delay)
        try:
            rows_out.append(
                extract_one_smart(
                    c,
                    browser=args.browser,
                    browser_fallback=args.browser_fallback,
                ),
            )
        except requests.RequestException as e:
            err = str(e)
            request_errors.append((c, err))
            print(f"{c}: REQUEST_ERROR: {err}", file=sys.stderr)
            rows_out.append(
                Row(
                    code=c,
                    url=DETAIL_URL + requests.utils.quote(c, safe=""),
                    birzhevoi_tovar=f"REQUEST_ERROR: {err}",
                    bazis_postavki=None,
                    data_posledney_sessii=None,
                    tsena_poslednyaya=None,
                    sessions_last_by_date={},
                ),
            )
        except RuntimeError as e:
            err = str(e)
            runtime_errors.append((c, err))
            print(f"{c}: RUNTIME_ERROR: {err}", file=sys.stderr)
            rows_out.append(
                Row(
                    code=c,
                    url=DETAIL_URL + requests.utils.quote(c, safe=""),
                    birzhevoi_tovar=f"RUNTIME_ERROR: {err}",
                    bazis_postavki=None,
                    data_posledney_sessii=None,
                    tsena_poslednyaya=None,
                    sessions_last_by_date={},
                ),
            )

    history_path = Path(args.history_output)
    history = _read_history_csv(history_path)
    history = _merge_history(history, rows_out)
    history_rows = _write_history_csv(history_path, history)

    out_path = Path(args.output)
    existing_dates, existing_rows = _read_existing_wide_csv(out_path)
    run_date = _today_msk_date_str()
    fieldnames, wide_rows = _build_wide_rows(
        rows_out,
        existing_rows=existing_rows,
        date_columns=existing_dates,
        run_date=run_date,
    )
    _write_wide_csv(out_path, fieldnames, wide_rows)

    hist_path = Path(args.histogram_output)
    histogram_ok = _draw_price_histogram(
        hist_path,
        date_columns=[c for c in fieldnames if _is_date_column(c)],
        rows_for_csv=wide_rows,
    )

    print(f"Записано строк: {len(rows_out)} -> {out_path}")
    print(f"История цен обновлена: {history_rows} записей -> {history_path}")
    if plt is None:
        print("Гистограмма не построена: установите matplotlib (pip install -r requirements.txt)")
    elif not histogram_ok:
        print("Гистограмма не построена: нет числовых значений цен для визуализации")
    else:
        print(f"Гистограмма сохранена: {hist_path}")

    report_path = Path(args.report_output)
    report_ok = _write_xlsx_report(report_path, history)
    if not report_ok:
        print("Excel-отчёт не сформирован: установите openpyxl (pip install -r requirements.txt)")
    else:
        print(f"Excel-отчёт сохранён: {report_path}")

    total_errors = len(request_errors) + len(runtime_errors)
    if total_errors:
        print(
            "Выгрузка завершена с ошибками: "
            f"request={len(request_errors)}, runtime={len(runtime_errors)}, total={total_errors}",
            file=sys.stderr,
        )
        failed_codes = ", ".join(code for code, _ in (request_errors + runtime_errors))
        if failed_codes:
            print(f"Коды с ошибками: {failed_codes}", file=sys.stderr)

    recipients = list(args.mail_to) if args.mail_to else _mail_recipients_from_env()
    if recipients:
        try:
            send_report_email(report_path, recipients, subject=args.mail_subject)
        except (OSError, ValueError, smtplib.SMTPException) as e:
            print(f"Отчёт записан, но отправка почты не удалась: {e}", file=sys.stderr)
            return 2
        print(f"Письмо с Excel-вложением отправлено: {', '.join(recipients)}")

    if total_errors:
        return 1
    return 0


if __name__ == "__main__":
    raise SystemExit(main(sys.argv[1:]))
